"""
Scraper SEIA - Experiencia SGA y Geobiota
Genera un único Excel con columna empresa_consultora.
Listo para correr con Windows Task Scheduler.
"""

import re
import time
import logging
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, parse_qs

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

OUTPUT_FILE = Path(__file__).parent / "Experiencia SGA-GEOBIOTA.xlsx"

EMPRESAS = [
    {"nombre": "SGA",      "ruts": ["723882"]},
    {"nombre": "Geobiota", "ruts": ["724662", "2132659649"]},
]

BASE_URL = "https://seia.sea.gob.cl"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sesión HTTP
# ---------------------------------------------------------------------------

session = requests.Session()
session.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
})


def get(url, **kwargs):
    """GET con reintentos simples."""
    for intento in range(3):
        try:
            resp = session.get(url, timeout=30, **kwargs)
            resp.raise_for_status()
            return resp
        except requests.RequestException as e:
            if intento == 2:
                raise
            log.warning("Reintento %d para %s — %s", intento + 1, url, e)
            time.sleep(2)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(text):
    if text is None:
        return None
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def normalizar_forma(texto):
    if not texto:
        return None
    t = texto.lower()
    if "estudio de impacto ambiental" in t:
        return "Estudio de Impacto Ambiental"
    if "declaraci" in t and "impacto ambiental" in t:
        return "Declaración de Impacto Ambiental"
    return None


# ---------------------------------------------------------------------------
# Scraping lista de proyectos (con paginación)
# ---------------------------------------------------------------------------

def scrape_lista_rut(rut: str) -> list[dict]:
    """Devuelve todos los proyectos de un RUT iterando páginas."""
    filas = []
    pagina = 1

    while True:
        url = (
            f"{BASE_URL}/busqueda/empresa_detalle.php"
            f"?rut_empresa={rut}&tipo=CONSULTORES&subtipo=EMPRESA&_paginador_refresh=1&_paginador_fila_actual={pagina}"
        )
        log.info("  Página %d — RUT %s", pagina, rut)
        resp = get(url)
        soup = BeautifulSoup(resp.text, "html.parser")

        tabla = soup.select_one("table.tabla")
        if not tabla:
            break

        filas_pagina = []
        for tr in tabla.find_all("tr")[1:]:
            tds = tr.find_all("td")
            if len(tds) < 7:
                continue
            a = tds[1].find("a", href=True)
            if not a:
                continue
            href = a["href"]
            id_exp = parse_qs(urlparse(href).query).get("id_expediente", [None])[0]
            filas_pagina.append({
                "n":               tds[0].get_text(strip=True),
                "nombre_proyecto": a.get_text(" ", strip=True),
                "id_expediente":   id_exp,
                "url_expediente":  href,
                "region":          tds[3].get_text(strip=True),
                "inversion_mmu":   tds[4].get_text(strip=True),
                "fecha":           tds[5].get_text(strip=True),
                "estado":          tds[6].get_text(strip=True),
            })

        if not filas_pagina:
            break

        # Detectar duplicados: si todos los IDs de esta página ya los tenemos, parar
        ids_actuales = {f["id_expediente"] for f in filas}
        ids_nuevos = [f for f in filas_pagina if f["id_expediente"] not in ids_actuales]
        if not ids_nuevos:
            log.info("  Página %d sin proyectos nuevos — fin de paginación", pagina)
            break

        filas.extend(ids_nuevos)
        pagina += 1
        time.sleep(0.5)

    return filas


# ---------------------------------------------------------------------------
# Scraping ficha individual
# ---------------------------------------------------------------------------

def scrape_ficha(id_expediente: str) -> dict:
    url = f"{BASE_URL}/expediente/ficha/fichaPrincipal.php?modo=ficha&id_expediente={id_expediente}"
    referer = f"{BASE_URL}/expediente/expedientesEvaluacion.php?modo=ficha&id_expediente={id_expediente}"
    resp = get(url, headers={"Referer": referer})
    soup = BeautifulSoup(resp.text, "html.parser")

    tipo_proyecto = None
    titular = None
    consultor = None

    # Tipo de proyecto
    for row in soup.select("div.row.gx-0.sg-row-file-data, div.row.sg-row-file-description"):
        cols = row.find_all("div", recursive=False)
        if len(cols) < 2:
            continue
        left  = clean(cols[0].get_text(" ", strip=True))
        right = clean(cols[1].get_text(" ", strip=True))
        if left and left.lower() == "tipo de proyecto":
            tipo_proyecto = right
            break

    # Titular y Consultor/a desde acordeones
    for btn in soup.select(".accordion-button"):
        titulo = clean(btn.get_text(" ", strip=True))
        target = btn.get("data-bs-target", "")
        if not target.startswith("#"):
            continue
        panel = soup.select_one(target)
        if not panel:
            continue
        nombre = None
        for row in panel.select("div.row.sg-row-file-description"):
            cols = row.find_all("div", recursive=False)
            if len(cols) < 2:
                continue
            left  = clean(cols[0].get_text(" ", strip=True))
            right = clean(cols[1].get_text(" ", strip=True))
            if left and left.lower() == "nombre":
                nombre = right
                break
        if titulo and titulo.lower() == "titular":
            titular = nombre
        elif titulo and titulo.lower() == "consultor/a":
            consultor = nombre

    # Forma de presentación
    forma = None
    for h in soup.select("h2.sg-subtitle"):
        forma = normalizar_forma(clean(h.get_text(" ", strip=True)))
        if forma:
            break

    # Descripción del proyecto
    descripcion = None
    for sp in soup.find_all("span"):
        label = clean(sp.get_text(" ", strip=True))
        if label and label.lower() == "descripción del proyecto":
            row = sp.find_parent("div", class_=lambda c: c and "row" in c.split())
            if row:
                sig = row.find_next_sibling("div")
                while sig:
                    dd = sig.select_one("div.sg-description-file")
                    if dd:
                        descripcion = clean(dd.get_text("\n", strip=True))
                        break
                    sig = sig.find_next_sibling("div")
            break
    if not descripcion:
        dd = soup.select_one("div.sg-description-file")
        if dd:
            descripcion = clean(dd.get_text("\n", strip=True))

    return {
        "id_expediente":      str(id_expediente),
        "tipo_proyecto_ficha": tipo_proyecto,
        "titular_ficha":      titular,
        "consultor_ficha":    consultor,
        "forma_presentacion": forma,
        "descripcion_proyecto": descripcion,
    }


# ---------------------------------------------------------------------------
# Formato Excel
# ---------------------------------------------------------------------------

def _aplicar_formato_tabla(path: Path, sheet_name: str = "Experiencia"):
    wb = load_workbook(path)
    ws = wb[sheet_name]

    # Rango de la tabla (todas las filas con datos)
    max_col_letter = get_column_letter(ws.max_column)
    rango = f"A1:{max_col_letter}{ws.max_row}"

    tabla = Table(displayName="TablaExperiencia", ref=rango)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",  # Azul estándar de Excel
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)

    # Ancho automático por columna
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        # Limitar columnas de texto largo (descripción) a 60 chars de ancho
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # Fila de encabezado fija al hacer scroll
    ws.freeze_panes = "A2"

    wb.save(path)


# ---------------------------------------------------------------------------
# Flujo principal
# ---------------------------------------------------------------------------

def main():
    log.info("=== Inicio actualización %s ===", datetime.now().strftime("%Y-%m-%d %H:%M"))
    todos = []

    for empresa in EMPRESAS:
        nombre = empresa["nombre"]
        log.info("Procesando %s...", nombre)
        filas = []

        for rut in empresa["ruts"]:
            nuevas = scrape_lista_rut(rut)
            log.info("  RUT %s → %d proyectos", rut, len(nuevas))
            filas.extend(nuevas)

        if not filas:
            log.warning("Sin proyectos para %s", nombre)
            continue

        df = pd.DataFrame(filas)
        df["empresa_consultora"] = nombre

        # Eliminar duplicados por id_expediente (puede solapar entre RUTs)
        df = df.drop_duplicates(subset="id_expediente")

        # Scraping de fichas
        log.info("Scrapeando fichas de %s (%d proyectos)...", nombre, len(df))
        fichas = []
        for id_exp in df["id_expediente"].dropna():
            try:
                fichas.append(scrape_ficha(str(id_exp)))
                log.info("  OK %s", id_exp)
            except Exception as e:
                log.error("  ERROR %s — %s", id_exp, e)
                fichas.append({
                    "id_expediente":       str(id_exp),
                    "tipo_proyecto_ficha": None,
                    "titular_ficha":       None,
                    "consultor_ficha":     None,
                    "forma_presentacion":  None,
                    "descripcion_proyecto": None,
                })
            time.sleep(0.3)  # pausa cortés con el servidor

        df_fichas = pd.DataFrame(fichas)
        df = df.merge(df_fichas, on="id_expediente", how="left")
        todos.append(df)

    if not todos:
        log.error("No se obtuvo ningún dato. Abortando.")
        return

    df_final = pd.concat(todos, ignore_index=True)

    # Orden de columnas: empresa_consultora antes de consultor_ficha
    col_order = [
        "n",
        "nombre_proyecto",
        "id_expediente",
        "url_expediente",
        "region",
        "inversion_mmu",
        "fecha",
        "estado",
        "forma_presentacion",
        "tipo_proyecto_ficha",
        "titular_ficha",
        "empresa_consultora",
        "consultor_ficha",
        "descripcion_proyecto",
    ]
    # Incluir cualquier columna extra que pueda aparecer en el futuro
    extras = [c for c in df_final.columns if c not in col_order]
    df_final = df_final[col_order + extras]

    # Guardar y aplicar formato de tabla Excel
    df_final.to_excel(OUTPUT_FILE, index=False, sheet_name="Experiencia")
    _aplicar_formato_tabla(OUTPUT_FILE, sheet_name="Experiencia")
    log.info("Excel guardado → %s", OUTPUT_FILE)
    log.info("Total proyectos: %d (SGA: %d, Geobiota: %d)",
             len(df_final),
             len(df_final[df_final.empresa_consultora == "SGA"]),
             len(df_final[df_final.empresa_consultora == "Geobiota"]))
    log.info("=== Fin ===")


if __name__ == "__main__":
    main()
